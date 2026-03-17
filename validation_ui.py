import os
import time
import re
import fitz  # PyMuPDF
from ocr_engine import extract_text
from PIL import Image, ImageEnhance, ImageFilter, ImageTk
import io
import tkinter as tk
from openpyxl import load_workbook
import shutil
import logging

import sys

if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

MAX_PAGES_VALIDATION = 3
DPI_LOW = 150
WIN_WIDTH = 1000
WIN_HEIGHT = 700

FOLDER_IN  = os.path.join(BASE_DIR, "1_Entrant_Deposer_PDF")
FOLDER_OUT = os.path.join(BASE_DIR, "2_Traite_Succes")
FOLDER_ERR = os.path.join(BASE_DIR, "3_Erreur_A_Verifier")

# Mêmes 4 candidats que main_watcher.py
_excel_candidates = [
    os.path.join(os.path.dirname(BASE_DIR), "\u00e9ch\u00e9ancier factures ventes", "Echeancier_cible.xlsx"),  # PROD
    os.path.join(BASE_DIR, "Echeancier_cible.xlsx"),
    os.path.join(os.path.dirname(BASE_DIR), "Echeancier_cible.xlsx"),
    os.path.join(BASE_DIR, "References", "Echeancier_cible.xlsx"),
]
EXCEL_FILE = next((p for p in _excel_candidates if os.path.exists(p)), _excel_candidates[0])

logging.basicConfig(filename=os.path.join(BASE_DIR, "workflow.log"), level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')

# OCR Engine (EasyOCR via ocr_engine.py)

def preprocess_image(img):
    img = img.convert('L')
    enhancer = ImageEnhance.Contrast(img)
    img = enhancer.enhance(2.0)
    img = img.filter(ImageFilter.SHARPEN)
    return img

def extract_text_and_first_page_image(pdf_path):
    texts = []
    first_page_img = None
    doc = fitz.open(pdf_path)
    try:
        max_pages = min(len(doc), MAX_PAGES_VALIDATION)
        for page_num in range(max_pages):
            page = doc.load_page(page_num)
            pix = page.get_pixmap(dpi=DPI_LOW) # lower DPI for UI preview
            pil_img = Image.open(io.BytesIO(pix.tobytes()))
            if page_num == 0:
                first_page_img = pil_img.copy()

            # OCR is done on 300dpi internally usually, but 150dpi here for speed during UI
            img_ocr = preprocess_image(pil_img)
            text = extract_text(img_ocr)
            texts.append(text)
    except Exception as e:
        logging.error(f"Erreur OCR sur {pdf_path}: {e}")
    finally:
        try:
            doc.close()
        except Exception as close_err:
            logging.warning(f"Failed to close doc: {close_err}")
        full_text = "\n".join(texts) + "\n" if texts else ""
    return full_text, first_page_img

def parse_invoice_text(text):
    data = {"num_facture": "", "client": "", "date_facture": "", "date_echeance": "", "montant_ttc": "", "session": ""}

    # TAU_2026-413
    m_facture = re.search(r'(?i)(TAU_\d{4}[\-\_]\d+)', text)
    if m_facture: data["num_facture"] = m_facture.group(1).strip()

    all_dates = re.findall(r'(\d{2}[/.\-]\d{2}[/.\-]\d{4})', text)
    if all_dates:
        data["date_facture"] = all_dates[0]

    m_echeance = re.search(r'(?i)(?:échéance|echeance|règlement|limite)\s*.*?(?:\d{2}[/.\-]\d{2}[/.\-]\d{4})', text)
    if m_echeance:
         dates = re.findall(r'(\d{2}[/.\-]\d{2}[/.\-]\d{4})', m_echeance.group(0))
         if dates: data["date_echeance"] = dates[-1]
    elif len(all_dates) > 1:
         data["date_echeance"] = all_dates[-1]

    m_session = re.search(r'(?i)Session(?:\s*du)?\s*(\d{2}[/.\-]\d{2}[/.\-]\d{4}\s*au\s*\d{2}[/.\-]\d{2}[/.\-]\d{4})', text)
    if m_session: data["session"] = m_session.group(1).strip()

    m_client = re.search(r'(?i)SOCIETE\s*([^\n\r]+)', text)
    if m_client: data["client"] = m_client.group(1).strip()

    m_ttc_explicit = re.search(r'(?i)(?:Total\s*TTC|TTC|Net\s*à\s*payer).*?([\d\s]+[,.]\d{2})(?:\s*€|\s*EUR)?', text)
    if m_ttc_explicit:
        data["montant_ttc"] = m_ttc_explicit.group(1).replace(' ', '').strip()
    else:
        amounts = re.findall(r'([\d\s]+[,.]\d{2})\s*(?:€|EUR)', text)
        if amounts:
            try:
                clean_amounts = [float(a.replace(' ', '').replace(',', '.')) for a in amounts]
                data["montant_ttc"] = f"{max(clean_amounts):.2f}".replace('.', ',')
            except ValueError:
                data["montant_ttc"] = amounts[-1].replace(' ', '').strip()

    return data

def check_duplicate(ws, num_facture):
    if not num_facture:
        return False
    for row in range(2, ws.max_row + 2):
        cell_val = ws[f"A{row}"].value
        if cell_val and str(cell_val).strip() == str(num_facture).strip():
            return True
    return False

def inject_to_excel(data):
    if not os.path.exists(EXCEL_FILE):
        return False

    try:
        wb = load_workbook(EXCEL_FILE)
        try:
            ws = wb["Ventes_Factures"]
        except KeyError:
            ws = wb.active
            logging.warning("Onglet 'Ventes_Factures' introuvable, utilisation de l'onglet actif")

        if check_duplicate(ws, data.get("num_facture")):
            wb.close()
            return "DUPLICATE"

        found_row = ws.max_row + 1
        for r in range(2, ws.max_row + 2):
            if not ws[f"A{r}"].value:
                found_row = r
                break

        ws[f"A{found_row}"] = data.get("num_facture", "")
        ws[f"B{found_row}"] = data.get("client", "")
        ws[f"C{found_row}"] = data.get("type_facture", "B2B")
        ws[f"E{found_row}"] = data.get("session", "")
        ws[f"F{found_row}"] = data.get("date_facture", "")
        ws[f"G{found_row}"] = data.get("date_echeance", "")

        mnt_raw = data.get("montant_ttc", "")
        if mnt_raw not in (None, ""):
            try:
                # Accepte float (depuis main) ou string avec virgule (saisie UI)
                mnt = float(str(mnt_raw).replace(',', '.').replace(' ', ''))
                ws[f"H{found_row}"] = mnt
            except (ValueError, TypeError):
                ws[f"H{found_row}"] = mnt_raw

        wb.save(EXCEL_FILE)
        wb.close()
        return "SUCCESS"
    except Exception as e:
        logging.error(f"Erreur d'injection Excel : {e}")
        return "ERROR"

class ValidationUI:
    def __init__(self, root_window, pdf_path, extracted_data, img_preview):
        self.root = root_window
        self.root.title("Validation Facture - Google Antigravity")
        self.root.geometry(f"{WIN_WIDTH}x{WIN_HEIGHT}")

        self.pdf_path = pdf_path
        self.data = extracted_data
        self.img_preview = img_preview
        self.status = "CANCEL" # Par défaut
        self.final_data = {}

        self.setup_ui()

    def setup_ui(self):
        # --- Panel Gauche (Image) ---
        self.left_panel = tk.Frame(self.root, width=500, bg="gray")
        self.left_panel.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        if self.img_preview:
            # Resize image to fit height
            h_ratio = WIN_HEIGHT / self.img_preview.height
            new_size = (int(self.img_preview.width * h_ratio), WIN_HEIGHT)
            img_resized = self.img_preview.resize(new_size, Image.LANCZOS)
            self.tk_img = ImageTk.PhotoImage(img_resized)

            lbl_img = tk.Label(self.left_panel, image=self.tk_img)
            lbl_img.pack(padx=10, pady=10)

        # --- Panel Droit (Formulaire) ---
        self.right_panel = tk.Frame(self.root, width=400, padx=20, pady=20)
        self.right_panel.pack(side=tk.RIGHT, fill=tk.Y)

        tk.Label(self.right_panel, text="Vérification des Données", font=("Arial", 16, "bold")).pack(pady=20)

        self.entries = {}
        fields = [
            ("N° Facture", "num_facture"),
            ("Client", "client"),
            ("Date Facture", "date_facture"),
            ("Date Échéance", "date_echeance"),
            ("Session", "session"),
            ("Montant TTC", "montant_ttc")
        ]

        for label_text, key in fields:
            frame = tk.Frame(self.right_panel)
            frame.pack(fill=tk.X, pady=5)

            lbl = tk.Label(frame, text=label_text, width=15, anchor="w")
            lbl.pack(side=tk.LEFT)

            ent_var = tk.StringVar(value=self.data.get(key, ""))
            ent = tk.Entry(frame, textvariable=ent_var, width=30)
            ent.pack(side=tk.LEFT, padx=10)

            # Highlight missing fields
            if not ent_var.get():
                ent.config(bg="lightcoral")

            self.entries[key] = ent_var

        btn_frame = tk.Frame(self.right_panel)
        btn_frame.pack(fill=tk.X, pady=40)

        tk.Button(btn_frame, text="Valider & Injecter", command=self.on_validate, bg="lightgreen", height=2).pack(side=tk.LEFT, expand=True, fill=tk.X, padx=5)
        tk.Button(btn_frame, text="Rejeter", command=self.on_reject, bg="salmon", height=2).pack(side=tk.RIGHT, expand=True, fill=tk.X, padx=5)

    def on_validate(self):
        self.final_data = {key: var.get() for key, var in self.entries.items()}
        self.status = "SUCCESS"
        self.root.destroy()

    def on_reject(self):
        self.status = "REJECT"
        self.root.destroy()

def process_with_ui(pdf_path, prefilled_data=None):
    """
    Ouvre l'UI de validation pour une page de facture.

    Args:
        pdf_path: Chemin vers le PDF source (pour l'aperçu image).
        prefilled_data: dict de données pré-extraites par main_watcher.py.
                        Si None, l'UI extrait elle-même (appel standalone).

    Returns:
        tuple (status, final_data)
            - status   : "SUCCESS" | "REJECT" | "CANCEL"
            - final_data : dict des données validées/corrigées par l'utilisateur
                          (None si rejet/cancel)
        Note: le déplacement du fichier PDF est géré par l'appelant (process_pdf),
        PAS par cette fonction. L'injection Excel est faite ici uniquement si
        appelée en mode standalone (__main__).
    """
    logging.info(f"[UI] Ouverture validation pour: {os.path.basename(pdf_path)}")

    # Récupération de l'image de prévisualisation (première page)
    _, extracted_img = extract_text_and_first_page_image(pdf_path)

    # Données à pré-remplir : priorité aux données passées en paramètre
    if prefilled_data is not None:
        # Convertir float → string pour affichage dans les champs texte
        display_data = dict(prefilled_data)
        mnt = display_data.get("montant_ttc")
        if isinstance(mnt, float):
            display_data["montant_ttc"] = f"{mnt:.2f}".replace('.', ',')
        # Convertir dates datetime.date → string
        import datetime as _dt
        for k in ("date_facture", "date_echeance"):
            v = display_data.get(k)
            if isinstance(v, _dt.date):
                display_data[k] = v.strftime("%d/%m/%Y")
        data = display_data
    else:
        # Mode standalone : extraire nous-mêmes
        text, _ = extract_text_and_first_page_image(pdf_path)
        data = parse_invoice_text(text)

    # Ouverture UI — bloquant jusqu'à fermeture
    root_val = tk.Tk()
    app = ValidationUI(root_val, pdf_path, data, extracted_img)
    root_val.mainloop()
    time.sleep(0.5)  # laisser les ressources Tk se libérer

    if app.status == "SUCCESS":
        logging.info(f"[UI] Validation confirmée pour {os.path.basename(pdf_path)}")
        # Restituer type_facture depuis prefilled_data (non affiché dans le formulaire)
        final = dict(app.final_data)
        if prefilled_data and "type_facture" in prefilled_data:
            final["type_facture"] = prefilled_data["type_facture"]
        return "SUCCESS", final
    else:
        logging.info(f"[UI] Rejet par l'utilisateur pour {os.path.basename(pdf_path)}")
        return "REJECT", None

if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        pdf_path = sys.argv[1]
    else:
        test_file = os.path.join(BASE_DIR, "References", "Factures ventes non réglées part.3.pdf")
        if not os.path.exists(test_file):
            print("Fichier de test introuvable.")
            sys.exit(1)
        pdf_path = os.path.join(FOLDER_IN, "Test_UI_3.pdf")
        shutil.copy(test_file, pdf_path)

    if not os.path.exists(pdf_path):
        print(f"Fichier introuvable : {pdf_path}")
        sys.exit(1)

    status, final_data = process_with_ui(pdf_path)
    if status == "SUCCESS" and final_data:
        result = inject_to_excel(final_data)
        print(f"Injection : {result}")
        dest = FOLDER_OUT if result == "SUCCESS" else FOLDER_ERR
        try:
            shutil.move(os.path.abspath(pdf_path), os.path.join(dest, os.path.basename(pdf_path)))
        except Exception as e:
            print(f"Erreur déplacement : {e}")
    else:
        print("Rejet ou annulation — aucune injection.")
