import os
import time
import queue
import traceback
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
import re
import fitz  # PyMuPDF
from PIL import Image, ImageEnhance, ImageFilter
import io
from openpyxl import load_workbook
import shutil
import logging
import json
import datetime
from validation_ui import process_with_ui
from ocr_engine import extract_text
from client_dictionary import match_client

# File d'attente thread-safe : watchdog → thread principal
# Tkinter DOIT s'exécuter depuis le thread principal (contrainte Windows).
# Le thread watchdog enfile les chemins PDF ; le thread principal les traite.
_pdf_queue = queue.Queue()
# Set thread-safe des fichiers déjà enfilés ou en cours de traitement
# (évite le double-enfilage watchdog qui déclenche parfois 2 événements)
import threading
_processing_lock = threading.Lock()
_files_seen = set()
try:
    from win10toast import ToastNotifier
    toaster = ToastNotifier()
except ImportError:
    toaster = None

import sys
import argparse

if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# ── Mode test (--dry-run) ──────────────────────────────────────
_parser = argparse.ArgumentParser(add_help=False)
_parser.add_argument("--dry-run", action="store_true",
                     help="Traite les PDFs sans écrire dans Excel ni déplacer les fichiers")
_args, _ = _parser.parse_known_args()
DRY_RUN = _args.dry_run

DPI_HIGH = 200

# Détection automatique du nommage des dossiers (deux conventions possibles)
def _find_folder(base, *candidates):
    for name in candidates:
        p = os.path.join(base, name)
        if os.path.isdir(p):
            return p
    # Aucun existant : créer le premier candidat
    p = os.path.join(base, candidates[0])
    os.makedirs(p, exist_ok=True)
    return p

FOLDER_IN     = _find_folder(BASE_DIR, "1_Entrant_Deposer_PDF", "Entrant")
FOLDER_OUT    = _find_folder(BASE_DIR, "2_Traite_Succes", "Traite")
FOLDER_ERR    = _find_folder(BASE_DIR, "3_Erreur_A_Verifier", "Erreur")
FOLDER_BACKUP = _find_folder(BASE_DIR, "backups")

# Recherche du fichier Excel dans plusieurs emplacements candidats
_excel_candidates = [
    os.path.join(os.path.dirname(BASE_DIR), "\u00e9ch\u00e9ancier factures ventes", "Echeancier_cible.xlsx"),  # z:\NZBG\échéanciers\échéancier factures ventes\ (PROD)
    os.path.join(BASE_DIR, "Echeancier_cible.xlsx"),                          # à côté de l'EXE
    os.path.join(os.path.dirname(BASE_DIR), "Echeancier_cible.xlsx"),          # dossier parent
    os.path.join(BASE_DIR, "References", "Echeancier_cible.xlsx"),             # sous-dossier References/
]
EXCEL_FILE = next((p for p in _excel_candidates if os.path.exists(p)), _excel_candidates[0])

# Log config
import hashlib
from logging.handlers import RotatingFileHandler as _RotatingFileHandler
_log_handler = _RotatingFileHandler(
    os.path.join(BASE_DIR, "workflow.log"),
    maxBytes=5 * 1024 * 1024,  # 5 MB
    backupCount=3,
    encoding='utf-8'
)
_log_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
_log_handler.setFormatter(_log_formatter)
_console_handler = logging.StreamHandler()
_console_handler.setFormatter(_log_formatter)
root_logger = logging.getLogger()
root_logger.addHandler(_log_handler)
root_logger.addHandler(_console_handler)
root_logger.setLevel(logging.INFO)

# OCR Engine (EasyOCR via ocr_engine.py)

# ─────────────────────────────────────────────────────────────
# MOTS-CLES POUBELLE pour filtrer l'entête Pôle Tauroentum
# ─────────────────────────────────────────────────────────────
JUNK_WORDS = [
    "tauroentum", "centre louis", "delacour", "la ciotat",
    "siret", "tél", "tel :", "tel.", "facture", "numéro",
    "description", "date", "n° client", "échéance", "echeance",
    "éditeur", "editeur", "enrichi", "courrier", "http",
    "rcs", "marseille", "paca", "cgr", "cgi", "tvA",
    "déclaration", "declaration", "agrément", "agrement",
    "exonéra", "exonera", "capital", "sarl au",
    "1 sur 1", "l sur 1", "sur 1",
    "generateur", "gst2015",
]

# ─────────────────────────────────────────────────────────────
# LIGNES PARASITES à supprimer du texte avant extraction montant
# (ex: "SARL au capital de 7500€" apparaît sur chaque page)
# ─────────────────────────────────────────────────────────────
FOOTER_PATTERNS = [
    r'(?i)capital\s+de\s+[\d\s.,]+[€ëé]',
    r'(?i)siret\s+n.*\d{5}',
    r'(?i)d[ée]claration\s+d.*activit',
    r'(?i)ne\s+vaut\s+pas\s+agr[ée]ment',
    r'(?i)exon[ée]ra.*TVA',
    r'(?i)article.*L?5362',
    r'(?i)article.*261',
    r'(?i)pr[ée]fet\s+de\s+la\s+r[ée]gion',
    r'(?i)restant\s+du',      # solde client ≠ montant TTC
    r'(?i)encaissement',      # ligne acompte/avoir
]


def preprocess_image(img):
    img = img.convert('L')
    enhancer = ImageEnhance.Contrast(img)
    img = enhancer.enhance(2.0)
    img = img.filter(ImageFilter.SHARPEN)
    return img


def extract_invoices_by_page(pdf_path):
    """
    Parcourt TOUTES les pages du PDF.
    Retourne une liste de (num_page, texte_ocr) pour les pages avec un TAU_.
    Pages sans TAU_ (convocations, programmes) ignorees.

    Stratégie en deux passes :
    1) Texte natif PyMuPDF (quasi instantané) — suffit pour PDFs éditeur web.
    2) EasyOCR uniquement si le texte natif est trop court ou ne contient pas TAU_
       sur AUCUNE page (PDF scanné).
    """
    invoices = []
    try:
        doc = fitz.open(pdf_path)
        total = len(doc)
        logging.info(f"  -> PDF de {total} pages, scan complet...")

        # Passe 1 : texte natif sur toutes les pages (rapide)
        native_texts = []
        any_tau_native = False
        for page_num in range(total):
            page = doc.load_page(page_num)
            native = page.get_text("text").strip()
            native_texts.append(native)
            if re.search(r'TAU_\d{4}[-_]\d+', native):
                any_tau_native = True

        for page_num in range(total):
            native = native_texts[page_num]
            native_has_tau = bool(re.search(r'TAU_\d{4}[-_]\d+', native))

            # Décision : utiliser texte natif ou OCR ?
            # - Si le PDF entier a au moins un TAU_ natif (PDF éditeur) :
            #     → utiliser natif pour toutes les pages (y compris les pages
            #       sans TAU_ pour éviter faux positifs OCR)
            # - Si aucun TAU_ natif (PDF scanné) : toujours OCR
            if any_tau_native:
                if native_has_tau:
                    logging.info(f"  -> Page {page_num + 1}/{total} (natif) — TAU_ détecté")
                    invoices.append((page_num + 1, native))
                else:
                    logging.info(f"  -> Page {page_num + 1}/{total} (natif) — pas de TAU_, ignorée")
            else:
                # PDF scanné : OCR nécessaire
                logging.info(f"  -> Page {page_num + 1}/{total} (OCR)...")
                page = doc.load_page(page_num)
                pix = page.get_pixmap(dpi=DPI_HIGH)
                pil_img = Image.open(io.BytesIO(pix.tobytes()))
                img_processed = preprocess_image(pil_img)
                text = extract_text(img_processed)
                if re.search(r'TAU_\d{4}[-_]\d+', text):
                    logging.info(f"  -> TAU_ détecté page {page_num + 1}")
                    invoices.append((page_num + 1, text))

        doc.close()
        logging.info(f"  -> {len(invoices)} page(s) avec facture detectee(s)"
                     + (" [mode natif]" if any_tau_native else " [mode OCR]"))
    except Exception as e:
        logging.error(f"Erreur OCR sur {pdf_path}: {e}")
    return invoices


def string_to_date(date_str):
    # Supporte "10/02/2026", "10 / 02 / 2026", "10.02.2026", etc.
    m = re.match(r'(\d{2})\s*[/.\-]\s*(\d{2})\s*[/.\-]\s*(\d{4})', date_str.strip())
    if m:
        try:
            dt = datetime.date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
            if 2000 <= dt.year <= 2100:
                return dt
        except Exception:
            pass
    return None


def clean_montant(m_str):
    s = str(m_str).strip()
    # Supprimer les symboles monétaires et espaces insécables
    s = s.replace('\xa0', ' ').replace('€', '').replace('EUR', '').replace('eur', '')
    # Correction artefacts OCR : lettres O/o lues à la place de 0 dans les décimales
    # ex: "465.OO" -> "465.00", "0.0o" -> "0.00"
    s = re.sub(r'(?<=[.,\d])[Oo]', '0', s)
    s = re.sub(r'[Oo](?=[.,\d])', '0', s)
    # Détecter le séparateur décimal réel et gérer les espaces-milliers
    # Cas "2 800,00" ou "2 800.00" : espace = séparateur milliers
    # Cas "2.800,00" : point = séparateur milliers, virgule = décimal
    # Cas "2,800.00" : virgule = séparateur milliers, point = décimal
    s = s.strip()
    if ',' in s and '.' in s:
        if s.rfind(',') > s.rfind('.'):
            # ex: "2.800,00" → point=milliers, virgule=décimal
            s = s.replace('.', '').replace(' ', '').replace(',', '.')
        else:
            # ex: "2,800.00" → virgule=milliers, point=décimal
            s = s.replace(',', '').replace(' ', '')
    elif ',' in s:
        # ex: "2 800,00" ou "2800,00" → virgule=décimal
        s = s.replace(' ', '').replace(',', '.')
    else:
        # ex: "2 800.00" ou "2800.00" → point=décimal
        s = s.replace(' ', '')
    try:
        return float(s)
    except Exception:
        return None


def _is_junk_line(line):
    """Retourne True si la ligne fait partie de l'entête Pôle Tauroentum ou est une adresse."""
    lower = line.lower()

    # Exception : une ligne entièrement en MAJUSCULES (≥4 lettres) est probablement
    # un nom de société → ne jamais la rejeter via les filtres génériques ci-dessous.
    # (On laisse quand même passer le filtre JUNK_WORDS qui est spécifique à l'émetteur.)
    letters_only = re.sub(r'[^A-Za-zÀ-ÿ]', '', line)
    is_all_caps = len(letters_only) >= 4 and letters_only == letters_only.upper()

    for junk in JUNK_WORDS:
        if junk in lower:
            return True

    if is_all_caps:
        # Pour une ligne tout-caps, on ne filtre que les adresses et codes postaux explicites
        if re.search(r'\d{5}', line):          # code postal seul = ville
            return True
        if re.search(r'\b(RUE|AVENUE|BOULEVARD|CHEMIN|ROUTE|ALLEE|IMPASSE)\b', line):
            return True
        return False  # Nom de société présumé → conserver

    # Code postal (5 chiffres + espace ou fin de ligne)
    if re.search(r'\d{5}[\s$]', line):
        return True
    # Adresse (rue, avenue, etc.)
    if re.search(r'\b(rue|avenue|boulevard|chemin|route|allée|impasse|zi|z\.i|za|z\.a|bp)\b', lower):
        return True
    # Ligne trop courte (<4 lettres utiles)
    if len(letters_only) < 4:
        return True
    # Date seule
    if re.match(r'^\d{2}[/.\-]\d{2}[/.\-]\d{4}', line.strip()):
        return True
    # Ligne de description de prestation (pas un nom de client)
    if re.search(r'\b(formation|session|durée|duree|stage|module|intitulé|intitule|'
                 r'prestation|programme|objectif|stagiaire)\b', lower):
        return True
    return False


def _extract_client_from_allcaps(lines, max_line=None):
    """
    Cherche dans `lines` (limité aux max_line premières si précisé)
    la première ligne entièrement en MAJUSCULES qui n'est pas du junk
    et qui est potentiellement un nom de société.
    Retourne le candidat ou None.
    """
    search_lines = lines[:max_line] if max_line else lines
    for line in search_lines:
        candidate = line.strip()
        if not candidate:
            continue
        letters = re.sub(r'[^A-Za-zÀ-ÿ]', '', candidate)
        if len(letters) < 3:
            continue
        # Doit être tout en majuscules
        if letters != letters.upper():
            continue
        # Ne pas prendre les mots-clés parasites de l'émetteur
        lower = candidate.lower()
        if any(junk in lower for junk in JUNK_WORDS):
            continue
        # Pas une adresse
        if re.search(r'\b(RUE|AVENUE|BOULEVARD|CHEMIN|ROUTE|ALLEE|IMPASSE)\b', candidate):
            continue
        if re.search(r'\d{5}', candidate):
            continue
        # Pas une date seule
        if re.match(r'^\d{2}[/.\-]\d{2}[/.\-]\d{4}', candidate.strip()):
            continue
        # Pas un numéro de facture
        if re.search(r'TAU_\d', candidate):
            continue
        return candidate
    return None


def parse_invoice_text(text):
    data = {
        "num_facture": "",
        "client": "",
        "date_facture": None,
        "date_echeance": None,
        "montant_ttc": None,
        "session": "",
        "type_facture": "B2B",
    }

    lines = [l.strip() for l in text.split('\n') if l.strip()]

    # ─── 1) N° Facture ──────────────────────────────────────────
    m_facture = re.search(r'(?i)(TAU_\d{4}[\-\_]\d+)', text)
    if m_facture:
        data["num_facture"] = m_facture.group(1).strip()

    # ─── 2) Dates ────────────────────────────────────────────────
    # Priorité : chercher la ligne contenant TAU_ qui a le format :
    # "TAU_XXXX-YYY  DATE_FACTURE  N_CLIENT  DATE_ECHEANCE"
    tau_line_dates = None
    for line in lines:
        if "TAU_" in line:
            ds = re.findall(r'\d{2}[\-/\.]\d{2}[\-/\.]\d{4}', line)
            if len(ds) >= 2:
                tau_line_dates = ds
                break
            elif len(ds) == 1:
                tau_line_dates = ds
                break

    if tau_line_dates and len(tau_line_dates) >= 2:
        data["date_facture"] = string_to_date(tau_line_dates[0])
        data["date_echeance"] = string_to_date(tau_line_dates[1])
    elif tau_line_dates and len(tau_line_dates) == 1:
        data["date_facture"] = string_to_date(tau_line_dates[0])
        if data["date_facture"]:
            data["date_echeance"] = data["date_facture"] + datetime.timedelta(days=30)
    else:
        # Fallback : première date valide du document
        valid_dates = []
        for d_str in re.findall(r'\d{2}\s*[/.\-]\s*\d{2}\s*[/.\-]\s*\d{4}', text):
            dt = string_to_date(d_str)
            if dt:
                valid_dates.append(dt)
        if valid_dates:
            data["date_facture"] = valid_dates[0]
            data["date_echeance"] = valid_dates[0] + datetime.timedelta(days=30)

    # ─── 3) Session ──────────────────────────────────────────────
    # Formats supportés :
    #   "Session du 16/03/2026 au 16/03/2026"
    #   "Session 16/03/2026 au 16/03/2026"
    #   "Session du 16.03.2026 au 16.03.2026"  (OCR peut mettre des points)
    #   Dates avec espaces OCR : "16 / 03 / 2026"
    _date_pat = r'\d{2}\s*[/.\-]\s*\d{2}\s*[/.\-]\s*\d{4}'
    m_session = re.search(
        r'(?i)Session\s+(?:du\s+)?(' + _date_pat + r')\s+au\s+(' + _date_pat + r')', text)
    if m_session:
        d1 = re.sub(r'\s', '', m_session.group(1))
        d2 = re.sub(r'\s', '', m_session.group(2))
        data["session"] = f"{d1} au {d2}"

    # ─── 4) TYPE (CPF / CDC / B2B) — détecté SUR TOUT LE TEXTE ─
    text_lower = text.lower()
    if re.search(r'compte\s+personnel\s+de\s+formation', text_lower):
        data["type_facture"] = "CPF"
    elif re.search(r'caisse\s+des\s+d.p.ts', text_lower):
        data["type_facture"] = "CDC"

    # ─── 5) CLIENT — extraction multi-passes ────────────────────
    client_name = None
    tau_index = -1
    for i, l in enumerate(lines):
        if "TAU_" in l:
            tau_index = i
            break

    # Passe 0 : fuzzy matching sur les lignes tout-MAJUSCULES du document
    # (le bloc destinataire OCR peut être n'importe où — cette passe l'attrape)
    # On limite au premier 40% du document pour éviter les lignes de bas de page.
    top_limit = max(tau_index + 5, len(lines) * 2 // 5) if tau_index >= 0 else len(lines) * 2 // 5
    allcaps_candidates = []
    for line in lines[:top_limit]:
        candidate = line.strip()
        letters = re.sub(r'[^A-Za-zÀ-ÿ]', '', candidate)
        if len(letters) < 3:
            continue
        if letters != letters.upper():
            continue
        if any(junk in candidate.lower() for junk in JUNK_WORDS):
            continue
        if re.search(r'\d{5}', candidate):
            continue
        if re.search(r'\b(RUE|AVENUE|BOULEVARD|CHEMIN|ROUTE|ALLEE|IMPASSE)\b', candidate):
            continue
        if re.match(r'^\d{2}[/.\-]\d{2}[/.\-]\d{4}', candidate.strip()):
            continue
        if re.search(r'TAU_\d', candidate):   # numéro de facture ≠ client
            continue
        allcaps_candidates.append(candidate)

    # Tenter un fuzzy match sur chaque candidat tout-caps : prendre le meilleur score ≥ 0.65
    best_caps_match = None
    best_caps_score = 0.0
    for candidate in allcaps_candidates:
        corrected, score, _ = match_client(candidate)
        if score >= 0.65 and score > best_caps_score:
            best_caps_score = score
            best_caps_match = (candidate, corrected, score)

    if best_caps_match:
        client_name = best_caps_match[0]  # conserve le nom brut pour la suite (fuzzy appliqué après)

    # Passe 1 : remonter les lignes au-dessus du TAU_
    if not client_name and tau_index > 0:
        for i in range(tau_index - 1, max(-1, tau_index - 10), -1):
            candidate = lines[i].strip()
            if _is_junk_line(candidate):
                continue
            candidate = re.sub(r'[_\(\)\[\]{}]', '', candidate).strip()
            candidate = re.sub(r'\beŸ\b', '', candidate).strip()
            if len(candidate) >= 3:
                client_name = candidate
                break

    # Fallback A: forme juridique sur les lignes avant TAU_
    if not client_name and tau_index > 0:
        for i in range(tau_index - 1, max(-1, tau_index - 15), -1):
            candidate = lines[i].strip()
            m_soc = re.search(
                r'([A-Z][A-Z0-9 &\-\.]{1,30}\s+(?:SARL|SAS|SA|EURL|SASU|ASSOCIATION|SNC|SCOP|GIE)\b'
                r'|(?:SARL|SAS|EURL|SASU|ASSOCIATION|SNC)\s+[A-Z][A-Z0-9 &\-\.]{2,30})',
                candidate)
            if m_soc:
                client_name = m_soc.group(0).strip().rstrip('&- ')
                if len(re.sub(r'[^A-Za-z]', '', client_name)) >= 4:
                    break
                client_name = None

    # Fallback B: forme juridique n'importe où dans le texte
    if not client_name:
        m_soc = re.search(r'\b((?:SARL|SAS|SA|EURL|SASU|ASSOCIATION)\s+[A-Z][A-Z &\-\.]+)', text)
        if m_soc:
            client_name = m_soc.group(1).strip()

    # Fallback C : première ligne tout-MAJUSCULES dans le premier tiers du document
    if not client_name:
        client_name = _extract_client_from_allcaps(lines, max_line=max(10, len(lines) // 3))

    # Si le type est CPF, le client est "CPF" (Caisse des Dépôts)
    if data["type_facture"] == "CPF":
        data["client"] = "CPF"
    elif data["type_facture"] == "CDC":
        data["client"] = "CDC"
    elif client_name:
        # Fuzzy matching avec le dictionnaire de clients connus
        corrected, score, was_corrected = match_client(client_name)
        if was_corrected:
            alpha_ratio = len(re.sub(r'[^A-Za-zÀ-ÿ]', '', client_name)) / max(len(client_name), 1)
            if alpha_ratio < 0.60:
                logging.warning(
                    f"Client corrigé SUSPECT (bruit OCR) : '{client_name}' -> '{corrected}' "
                    f"(score={score:.2f}, alpha_ratio={alpha_ratio:.0%}) — validation UI requise")
                data["_client_noise"] = True
            else:
                logging.info(f"Client corrige: '{client_name}' -> '{corrected}' (score={score:.2f})")
        data["client"] = corrected

    # ─── 6) MONTANT TTC — Extraction en 4 passes ────────────────
    montant_val = None

    # NETTOYAGE : Supprimer les lignes de footer avant extraction
    # pour éviter que "capital de 7500€" pollue les résultats
    clean_lines = []
    for line in lines:
        is_footer = False
        for pattern in FOOTER_PATTERNS:
            if re.search(pattern, line):
                is_footer = True
                break
        if not is_footer:
            clean_lines.append(line)
    clean_text = '\n'.join(clean_lines)
    # Corriger les O/o OCR dans les montants du texte nettoyé (ex: "465.OO€" -> "465.00€")
    clean_text = re.sub(r'(\d)[Oo]([Oo€\s])', lambda m: m.group(0).replace('O','0').replace('o','0'), clean_text)
    clean_text = re.sub(r'(\d\.[Oo0][Oo])', lambda m: m.group(0).replace('O','0').replace('o','0'), clean_text)

    # Passe 1 : "Total TTC" / "Net à payer" / "Montant TTC" / "Tarif" explicite
    # Pattern de nombre : gère "2 800,00", "2.800,00", "2800.00", "465,OO", etc.
    _num_pat = r'[\d\s]{1,10}[.,][\d\s]{1,3}(?:[.,]\d{1,2})?'
    m1 = re.search(
        r'(?i)(?:Total\s*TTC|Net\s*[àa]\s*payer|Montant\s+TTC|Tarif|^Montant)\s*[:\-]?\s*(' + _num_pat + r')\s*(?:€|EUR|[eë])?',
        clean_text, re.MULTILINE)
    if m1:
        montant_val = clean_montant(m1.group(1))

    # Passe 2 : Nombres suivis de "€" ou "EUR" n'importe où (hors footer)
    if not montant_val:
        amounts = []
        for m2 in re.finditer(r'([\d\s.,]+)\s*(?:€|EUR)', clean_text):
            v = clean_montant(m2.group(1))
            if v is not None and v > 10:
                amounts.append(v)
        if amounts:
            montant_val = max(amounts)

    # Passe 3 : Nombres comme "2800€" mal lus comme "2800ë"
    if not montant_val:
        amounts = []
        for m3 in re.finditer(r'(\d{2,})\s*[€ëé]', clean_text):
            v = clean_montant(m3.group(1))
            if v is not None and v > 50:
                amounts.append(v)
        if amounts:
            montant_val = max(amounts)

    # Passe 4 : Dernier recours — le plus gros nombre décimal dans le corps
    if not montant_val:
        amounts = []
        for line in clean_lines[-15:]:
            for m4 in re.findall(r'(\d{2,}[,.]\d{2})', line):
                v = clean_montant(m4)
                if v is not None and v > 50:
                    amounts.append(v)
        if amounts:
            montant_val = max(amounts)

    data["montant_ttc"] = montant_val

    # Détection montant aberrant : < 10€ sur une facture de formation est un artefact OCR
    # (SIRET, numéro de page, date partiellement lue comme nombre)
    if montant_val is not None and montant_val < 10.0:
        logging.warning(
            f"Montant suspect détecté : {montant_val}€ (< 10€) — validation UI requise")
        data["_montant_suspect"] = True

    return data


# ── Hashes PDF déjà traités (persistés entre les runs) ────────
_HASHES_FILE = os.path.join(BASE_DIR, "processed_hashes.txt")

def _load_hashes():
    if not os.path.exists(_HASHES_FILE):
        return set()
    with open(_HASHES_FILE, "r", encoding="utf-8") as f:
        return {line.strip() for line in f if line.strip()}

def _save_hash(h):
    with open(_HASHES_FILE, "a", encoding="utf-8") as f:
        f.write(h + "\n")

def _pdf_hash(filepath):
    h = hashlib.md5()
    with open(filepath, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()

_processed_hashes = _load_hashes()


# ── Backup Excel horodaté ──────────────────────────────────────
def _backup_excel():
    if not os.path.exists(EXCEL_FILE):
        return
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = os.path.join(FOLDER_BACKUP, f"Echeancier_cible_{ts}.xlsx")
    shutil.copy2(EXCEL_FILE, dest)
    logging.info(f"Backup Excel créé : {os.path.basename(dest)}")


def check_duplicate(ws, num_facture):
    if not num_facture:
        return False
    for row in range(2, ws.max_row + 2):
        cell_val = ws[f"A{row}"].value
        if cell_val and str(cell_val).strip().upper() == str(num_facture).strip().upper():
            return True
    return False


def inject_to_excel(data):
    if DRY_RUN:
        logging.info(f"  [DRY-RUN] Injection simulée : {data.get('num_facture')} / {data.get('client')} / {data.get('montant_ttc')}€")
        return "SUCCESS"

    if not os.path.exists(EXCEL_FILE):
        logging.error("Fichier Excel introuvable.")
        return "ERROR"

    _backup_excel()

    try:
        wb = load_workbook(EXCEL_FILE)
        if "Ventes_Factures" not in wb.sheetnames:
            logging.error(
                "Onglet 'Ventes_Factures' introuvable dans le classeur Excel. "
                "Vérifier le fichier et renommer l'onglet correctement.")
            wb.close()
            return "ERROR"
        ws = wb["Ventes_Factures"]

        num_facture_val = data.get("num_facture")
        if check_duplicate(ws, num_facture_val):
            logging.warning(f"Facture {num_facture_val} déjà présente - ignorée.")
            wb.close()
            return "DUPLICATE"

        # Trouver la première ligne vide
        target_row = ws.max_row + 1
        for r in range(2, ws.max_row + 2):
            if not ws[f"A{r}"].value:
                target_row = r
                break

        ws[f"A{target_row}"] = num_facture_val
        ws[f"B{target_row}"] = data.get("client", "")
        ws[f"C{target_row}"] = data.get("type_facture", "B2B")
        ws[f"E{target_row}"] = data.get("session", "")

        dfact = data.get("date_facture")
        if dfact:
            if isinstance(dfact, str):
                dfact = string_to_date(dfact)
            if isinstance(dfact, datetime.date):
                ws[f"F{target_row}"] = dfact
                ws[f"F{target_row}"].number_format = 'DD/MM/YYYY'

        dech = data.get("date_echeance")
        if dech:
            if isinstance(dech, str):
                dech = string_to_date(dech)
            if isinstance(dech, datetime.date):
                ws[f"G{target_row}"] = dech
                ws[f"G{target_row}"].number_format = 'DD/MM/YYYY'

        mnt = data.get("montant_ttc")
        if mnt is not None:
            ws[f"H{target_row}"] = float(mnt)

        wb.save(EXCEL_FILE)
        wb.close()
        return "SUCCESS"
    except Exception as e:
        logging.error(f"Erreur d'injection Excel : {e}")
        return "ERROR"


class InvoiceHandler(FileSystemEventHandler):
    def on_created(self, event):
        if event.is_directory:
            return
        filepath = event.src_path
        if filepath.lower().endswith(".pdf"):
            time.sleep(2)
            if not os.path.exists(filepath):
                return
            # Éviter double-enfilage (watchdog déclenche parfois deux événements)
            with _processing_lock:
                if filepath in _files_seen:
                    return
                _files_seen.add(filepath)
            logging.info(f"Nouveau fichier détecté (enfilé) : {filepath}")
            _pdf_queue.put(filepath)


def log_to_json(filepath, data, score, status):
    log_file = os.path.join(BASE_DIR, "workflow.json")

    # Safe dump for datetime objects
    safe_data = {}
    for k, v in data.items():
        if isinstance(v, datetime.date):
            safe_data[k] = v.isoformat()
        else:
            safe_data[k] = v

    entry = {
        "timestamp": datetime.datetime.now().isoformat(),
        "file": os.path.basename(filepath),
        "data_extracted": safe_data,
        "confidence_score": score,
        "status": status,
        "user": os.getlogin()
    }
    try:
        logs_list = []
        if os.path.exists(log_file):
            with open(log_file, "r", encoding="utf-8") as f:
                try:
                    logs_list = json.load(f)
                except ValueError:
                    logs_list = []
        logs_list.append(entry)
        with open(log_file, "w", encoding="utf-8") as f:
            json.dump(logs_list, f, indent=4, ensure_ascii=False)
    except Exception as e:
        logging.error(f"Erreur ecriture log JSON : {e}")


def calculate_confidence(data):
    conf_score = 10
    if not data.get("num_facture"):
        conf_score -= 3
    if not data.get("client"):
        conf_score -= 2
    if not data.get("date_facture"):
        conf_score -= 2
    if not data.get("date_echeance"):
        conf_score -= 2
    if not data.get("montant_ttc"):
        conf_score -= 3
    return max(0, conf_score)


def process_pdf(filepath):
    filename = os.path.basename(filepath)
    try:
        # ── Vérification hash (doublon inter-sessions) ─────────
        pdf_hash = _pdf_hash(filepath)
        if pdf_hash in _processed_hashes:
            logging.warning(f"PDF déjà traité (hash connu), ignoré : {filename}")
            return

        # ── Copie de travail dans un dossier temporaire ────────
        tmp_dir = os.path.join(BASE_DIR, "_tmp_work")
        os.makedirs(tmp_dir, exist_ok=True)
        work_path = os.path.join(tmp_dir, filename)
        shutil.copy2(filepath, work_path)
        logging.info(f"PDF détecté : {filename}")
        logging.info(f"Traitement en cours... (copie de travail : {work_path})")

        logging.info(f"Extraction du texte pour {filename}...")
        invoice_pages = extract_invoices_by_page(work_path)

        if not invoice_pages:
            logging.warning(f"Aucune facture TAU_ détectée dans {filename} - fichier ignoré")
            shutil.move(filepath, os.path.join(FOLDER_ERR, filename))
            return

        success_count = 0
        error_count = 0
        skip_count = 0

        for page_num, text in invoice_pages:
            data = parse_invoice_text(text)
            score = calculate_confidence(data)
            num = data.get('num_facture', f'page_{page_num}')
            logging.info(
                f"  [p.{page_num}] {num} | client={data.get('client', '?')} | "
                f"montant={data.get('montant_ttc', '?')} | score={score}/10")

            if data.get("type_facture") == "CPF":
                # Factures CPF : injection directe sans validation UI
                # (client = "CPF", données standardisées)
                logging.info(f"  [p.{page_num}] Facture CPF — injection directe (sans UI)")
                status = inject_to_excel(data)
            elif (score < 7
                  or data.get("_client_noise")
                  or data.get("_montant_suspect")
                  or data.get("montant_ttc") is None):
                # Score insuffisant, bruit OCR client, montant aberrant, ou montant absent
                if data.get("_client_noise"):
                    reason = "bruit OCR client"
                elif data.get("_montant_suspect"):
                    reason = f"montant suspect ({data.get('montant_ttc')}€ < 10€)"
                elif data.get("montant_ttc") is None:
                    reason = "montant absent"
                else:
                    reason = f"score {score}/10 < 7"
                logging.info(
                    f"  [p.{page_num}] {reason} — envoi vers UI de validation")
                ui_status, ui_data = process_with_ui(filepath, prefilled_data=data, page_num=page_num - 1)
                if ui_status == "SUCCESS" and ui_data:
                    # L'utilisateur a corrigé les données — on injecte ses corrections
                    status = inject_to_excel(ui_data)
                    data = ui_data  # pour log_to_json
                else:
                    # Rejet ou fermeture fenêtre — on skip cette page
                    logging.info(f"  [p.{page_num}] Rejet UI — page ignorée")
                    skip_count += 1
                    log_to_json(filepath, data, score, "REJECTED_BY_USER")
                    continue
            else:
                status = inject_to_excel(data)

            if status == "SUCCESS":
                success_count += 1
                log_to_json(filepath, data, score, "SUCCESS")
            elif status == "DUPLICATE":
                skip_count += 1
                log_to_json(filepath, data, score, "DUPLICATE")
            else:
                error_count += 1
                log_to_json(filepath, data, score, "ERROR")

        # ── Résumé final ───────────────────────────────────────
        logging.info(
            f"Traitement terminé : {success_count} succès / {error_count} erreurs"
            + (f" / {skip_count} ignorés" if skip_count else "")
            + (" [DRY-RUN]" if DRY_RUN else ""))

        # Enregistrer le hash pour éviter un retraitement futur
        if not DRY_RUN and error_count == 0:
            _processed_hashes.add(pdf_hash)
            _save_hash(pdf_hash)

        # Nettoyage copie de travail
        try:
            os.remove(work_path)
        except Exception:
            pass

        if toaster:
            try:
                pass  # win10toast désactivé (crash silencieux en EXE)
            except Exception as e:
                logging.warning(f"Erreur notification : {e}")

        # Déplacement du fichier original
        if not DRY_RUN:
            if error_count == 0:
                shutil.move(filepath, os.path.join(FOLDER_OUT, filename))
                logging.info(f"Fichier déplacé vers Traité : {filename}")
            else:
                shutil.move(filepath, os.path.join(FOLDER_ERR, filename))
                logging.info(f"Fichier déplacé vers Erreur : {filename}")

    except Exception as e:
        logging.error(f"Erreur globale sur {filename} : {e}")
        try:
            if not DRY_RUN:
                shutil.move(filepath, os.path.join(FOLDER_ERR, filename))
        except Exception as mv_err:
            logging.error(f"Failed to move file to error folder: {mv_err}")


def process_existing_files():
    """Enfile les PDFs déjà présents dans FOLDER_IN au démarrage."""
    existing = [f for f in os.listdir(FOLDER_IN) if f.lower().endswith('.pdf')]
    if existing:
        logging.info(f"Scan démarrage : {len(existing)} PDF(s) déjà présents dans {FOLDER_IN}")
        for filename in existing:
            filepath = os.path.join(FOLDER_IN, filename)
            with _processing_lock:
                _files_seen.add(filepath)
            logging.info(f"Enfilé au démarrage : {filename}")
            _pdf_queue.put(filepath)
    else:
        logging.info(f"Scan démarrage : aucun PDF en attente dans {FOLDER_IN}")


def start_watcher():
    """
    Lance watchdog dans un thread daemon, puis tourne en boucle principale
    pour dépiler _pdf_queue et appeler process_pdf.
    Tkinter DOIT s'exécuter depuis le thread principal → cette boucle tourne
    dans le thread principal, watchdog dans un thread séparé.
    """
    if DRY_RUN:
        logging.info("*** MODE DRY-RUN activé — aucune écriture Excel ni déplacement de fichiers ***")
    logging.info(f"Démarrage surveillance sur {FOLDER_IN}")
    print(f"Watchdog actif. Déposez un PDF dans {FOLDER_IN}")

    # Watchdog dans un thread daemon (ne bloque pas le thread principal)
    event_handler = InvoiceHandler()
    observer = Observer()
    observer.schedule(event_handler, FOLDER_IN, recursive=False)
    observer.start()

    # Enfile les PDFs déjà présents
    process_existing_files()

    # Boucle principale — dépile la queue et traite les PDFs
    try:
        while True:
            try:
                filepath = _pdf_queue.get(timeout=1)
            except queue.Empty:
                continue
            try:
                process_pdf(filepath)
            except Exception as e:
                logging.error(f"Erreur non gérée dans process_pdf : {e}\n{traceback.format_exc()}")
            finally:
                # Libérer le fichier du Set : permettre un retraitement manuel futur
                with _processing_lock:
                    _files_seen.discard(filepath)
    except KeyboardInterrupt:
        observer.stop()
    observer.join()


if __name__ == "__main__":
    start_watcher()
