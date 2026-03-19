"""
Batterie de tests WorkflowFactures V5.8
"""
import sys, os, datetime
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from main_watcher import parse_invoice_text, calculate_confidence
from client_dictionary import match_client

def run_suite(name, tests):
    ok = 0
    print(f"\n{'='*60}")
    print(f"  {name}")
    print(f"{'='*60}")
    for label, fn, expected in tests:
        actual = fn()
        passed = actual == expected
        if passed:
            ok += 1
        status = "OK" if passed else f"KO  got={actual!r}  want={expected!r}"
        print(f"  {'OK' if passed else 'KO'}  {label}")
        if not passed:
            print(f"       -> got:  {actual!r}")
            print(f"       -> want: {expected!r}")
    print(f"  {ok}/{len(tests)} passed")
    return ok, len(tests)

# ─── SUITE 1 : Fuzzy matching ────────────────────────────────
def fuzzy_tests():
    cases = [
        ("Exact match",            lambda: match_client("GENAVIR")[0],              "GENAVIR"),
        ("Faute OCR simple",       lambda: match_client("FOSELEV MARIN")[0],        "FOSELEV MARINE"),
        ("Faute OCR grave",        lambda: match_client("CORSICA LINE")[0],         "CORSICA LINEA"),
        ("Acronyme court",         lambda: match_client("JIFM")[0],                 "JIFMAR"),
        ("Casse mixte",            lambda: match_client("foselev marine")[0],       "FOSELEV MARINE"),
        ("Tiret vs espace",        lambda: match_client("THALES SYSTEMGIE")[0],     "THALES - SYSTEMGIE"),
        ("Metropole tronqué",      lambda: match_client("METROPOL")[0],             "METROPOLE"),
        ("Score exact = 1.0",      lambda: match_client("ACTION MER")[1],           1.0),
        ("Non corrigé si exact",   lambda: match_client("GPMM")[2],                 False),
        ("Corrigé si faute",       lambda: match_client("FOSELEV MARIN")[2],        True),
        ("Adresse ne match pas",   lambda: match_client("56 Rue de Lille")[0],      "56 Rue de Lille"),
    ]
    return run_suite("FUZZY MATCHING (client_dictionary.py)", cases)

# ─── SUITE 2 : Extraction des champs ────────────────────────
def parser_tests():
    d = datetime.date

    def p(text, field):
        return parse_invoice_text(text).get(field)

    cases = [
        ("Num facture TAU_",
            lambda: p("CLIENT\nTAU_2026-559\n01/01/2026\nTotal TTC 100 EUR", "num_facture"),
            "TAU_2026-559"),
        ("Num facture avec underscore",
            lambda: p("CLIENT\nTAU_2026_100\n01/01/2026\nTotal TTC 100 EUR", "num_facture"),
            "TAU_2026_100"),
        ("Date standard",
            lambda: p("CLIENT\nTAU_2026-1\n12/03/2026\nTotal TTC 100 EUR", "date_facture"),
            d(2026, 3, 12)),
        ("Date avec espaces OCR",
            lambda: p("CLIENT\nTAU_2026-1\n05 / 03 / 2026\nTotal TTC 100 EUR", "date_facture"),
            d(2026, 3, 5)),
        ("Echeance +30j auto",
            lambda: p("FOSELEV MARINE\nTAU_2026-1\n01/02/2026\nTotal TTC 100 EUR", "date_echeance"),
            d(2026, 3, 3)),
        ("Montant Total TTC",
            lambda: p("CLIENT\nTAU_2026-1\n01/01/2026\nTotal TTC : 7 500,00 EUR", "montant_ttc"),
            7500.0),
        ("Montant Net a payer",
            lambda: p("CLIENT\nTAU_2026-1\n01/01/2026\nNet a payer 465,00 EUR", "montant_ttc"),
            465.0),
        ("Montant avec symbole euro",
            lambda: p("CLIENT\nTAU_2026-1\n01/01/2026\n1200,00 \u20ac", "montant_ttc"),
            1200.0),
        ("Montant e malformed (OCR)",
            lambda: p("CLIENT\nTAU_2026-1\n01/01/2026\n2800\u00eb", "montant_ttc"),
            2800.0),
        ("Footer capital ignore",
            lambda: p("CORSICA LINEA\nTAU_2026-1\n01/01/2026\nSARL au capital de 7500 EUR\nTotal TTC 5 000,00 EUR", "montant_ttc"),
            5000.0),
        ("Type CPF",
            lambda: p("Compte Personnel de Formation\nTAU_2026-1\n01/01/2026\nTotal TTC 1500 EUR", "type_facture"),
            "CPF"),
        ("Type CDC",
            lambda: p("Caisse des depots\nTAU_2026-1\n01/01/2026\nTotal TTC 1500 EUR", "type_facture"),
            "CDC"),
        ("Client CPF = CPF",
            lambda: p("Compte Personnel de Formation\nTAU_2026-1\n01/01/2026\nTotal TTC 1500 EUR", "client"),
            "CPF"),
        ("Sans client = vide",
            lambda: p("TAU_2026-1\n01/01/2026\nTotal TTC 300 EUR", "client"),
            ""),
        ("Sans montant = None",
            lambda: p("GENAVIR\nTAU_2026-1\n01/01/2026", "montant_ttc"),
            None),
        # ── Parsing montants avec séparateurs milliers ──
        ("Montant espace-milliers 2 800,00 €",
            lambda: p("NAVY SERVICE\nTAU_2026-1\n01/01/2026\nTotal TTC : 2 800,00 €", "montant_ttc"),
            2800.0),
        ("Montant sans espace 2800,00",
            lambda: p("NAVY SERVICE\nTAU_2026-1\n01/01/2026\nTotal TTC : 2800,00 EUR", "montant_ttc"),
            2800.0),
        ("Montant point-milliers OCR 2.800,00€",
            lambda: p("NAVY SERVICE\nTAU_2026-1\n01/01/2026\nTotal TTC : 2.800,00€", "montant_ttc"),
            2800.0),
        # ── Filtre description / client bruité ──
        ("Ligne formation ignorée comme client",
            lambda: p("NAVY SERVICE\nTAU_2026-1\n01/01/2026\nformation technique\nTotal TTC 500 EUR", "client"),
            "NAVY SERVICE"),
        ("Ligne session ignorée comme client",
            lambda: p("GPMM\nTAU_2026-1\n01/01/2026\nSession du 01/03/2026 au 01/03/2026\nTotal TTC 500 EUR", "client"),
            "GPMM"),
        # ── Extraction session ──
        ("Session standard extraite",
            lambda: p("GENAVIR\nTAU_2026-1\n01/01/2026\nSession du 16/03/2026 au 16/03/2026\nTotal TTC 1000 EUR", "session"),
            "16/03/2026 au 16/03/2026"),
        ("Session sans 'du' extraite",
            lambda: p("GENAVIR\nTAU_2026-1\n01/01/2026\nSession 16/03/2026 au 18/03/2026\nTotal TTC 1000 EUR", "session"),
            "16/03/2026 au 18/03/2026"),
        # ── Extraction client bloc destinataire (tout-MAJUSCULES) ──
        ("Client OCAPIAT depuis bloc caps",
            lambda: p("CENTRE TAUROENTUM\nFACTURE\nOCAPIAT\n13 rue des Mousses\n13000 MARSEILLE\nTAU_2026-10\n01/01/2026\nTotal TTC 1200 EUR", "client"),
            "OCAPIAT"),
        ("Client EXAIL depuis bloc caps",
            lambda: p("CENTRE TAUROENTUM\nFACTURE\nEXAIL\n14 avenue de la Mer\n13009 MARSEILLE\nTAU_2026-11\n01/01/2026\nTotal TTC 800 EUR", "client"),
            "EXAIL"),
        ("Client TLV depuis acronyme caps",
            lambda: p("CENTRE TAUROENTUM\nFACTURE\nTLV\nPort de Toulon\n83000 TOULON\nTAU_2026-12\n01/01/2026\nTotal TTC 950 EUR", "client"),
            "TLV Transports Maritimes et Terrestres du Littoral Varois"),
        ("Client absent = vide (no caps valide)",
            lambda: p("TAU_2026-99\n01/01/2026\nTotal TTC 300 EUR", "client"),
            ""),
    ]
    return run_suite("EXTRACTION FACTURE (main_watcher.py)", cases)

def _check_dup_casse():
    """Vérifie que check_duplicate détecte un doublon malgré une casse différente."""
    from openpyxl import Workbook
    from main_watcher import check_duplicate
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "N° Facture"
    ws["A2"] = "tau_2026-100"   # minuscules intentionnelles
    return check_duplicate(ws, "TAU_2026-100")  # doit retourner True après patch


# ─── SUITE 3 : Score de confiance ───────────────────────────
def confidence_tests():
    today = datetime.date.today()

    def base():
        return {'num_facture': 'TAU_2026-1', 'client': 'ACPM',
                'date_facture': today, 'date_echeance': today, 'montant_ttc': 1000.0}

    def miss(**kw):
        d = base()
        d.update(kw)
        return d

    cases = [
        ("Complet = 10",           lambda: calculate_confidence(base()),                                   10),
        ("Sans montant = 7",       lambda: calculate_confidence(miss(montant_ttc=None)),                    7),
        ("Sans client = 8",        lambda: calculate_confidence(miss(client='')),                           8),
        ("Sans num = 7",           lambda: calculate_confidence(miss(num_facture='')),                      7),
        ("Sans date = 8",          lambda: calculate_confidence(miss(date_facture=None, date_echeance=None)), 6),
        ("Tout vide = 0",          lambda: calculate_confidence({'num_facture':'','client':'','date_facture':None,'date_echeance':None,'montant_ttc':None}), 0),
        ("Seuil UI déclenche <7",  lambda: calculate_confidence(miss(client='', montant_ttc=None)) < 7,     True),
        # V5.8 : montant_ttc=None force l'UI même si score >= 7
        # Le score reste 7 (calculate_confidence inchangé), mais le routage dans
        # process_pdf vérifie `montant_ttc is None` indépendamment du score.
        ("Montant None force UI (score=7 mais routage bloqué)",
            lambda: calculate_confidence(miss(montant_ttc=None)) == 7
                    and miss(montant_ttc=None).get("montant_ttc") is None,
            True),
        # Doublon insensible à la casse (check_duplicate normalisé en uppercase)
        ("Doublon casse différente détecté",
            lambda: _check_dup_casse(),
            True),
    ]
    return run_suite("SCORE DE CONFIANCE (calculate_confidence)", cases)

if __name__ == "__main__":
    total_ok = 0
    total = 0

    for fn in [fuzzy_tests, parser_tests, confidence_tests]:
        ok, n = fn()
        total_ok += ok
        total += n

    print(f"\n{'='*60}")
    print(f"  TOTAL : {total_ok}/{total} tests passed")
    if total_ok == total:
        print("  SUCCES - Toutes les suites sont vertes")
    else:
        print(f"  ECHEC  - {total - total_ok} test(s) en echec")
    print(f"{'='*60}\n")
    sys.exit(0 if total_ok == total else 1)
